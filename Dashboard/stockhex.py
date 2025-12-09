import tkinter as tk
from tkinter import ttk, messagebox
from stockhex_db import (
    init_db, add_product, get_all_products, get_products_by_type,
    update_product, delete_product
)
import openpyxl
from openpyxl.styles import Font, Alignment
import os

init_db()

class ProductApp:
    COLUMNS = {
        "bolt": ("ID", "Spanner / Size", "Type", "Length / mm", "Unit", "Price", "Grade", "Thread", "Place"),
        "nut": ("ID", "Spanner / Size", "Type", "Unit", "Price", "Grade", "Thread", "Place"),
        "washer": ("ID", "Spanner / Size", "Type", "Unit", "Price", "Quality", "Place"),
        "default": ("ID", "Spanner / Size", "Type", "Length / mm", "Unit", "Price", "Grade", "Thread", "Place"),
    }

    def __init__(self, root):
        self.root = root
        self.root.title("Hexagon Manager")
        self.root.geometry("1200x650")
        self.root.configure(bg="#f5f8ff")  # GPay-inspired background

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TLabel", background="#f5f8ff", font=("Helvetica", 11))
        style.configure("TButton", background="#1a73e8", foreground="white", font=("Helvetica", 11, "bold"), padding=5)
        style.map("TButton", background=[("active", "#1669c1")])
        style.configure("TEntry", font=("Helvetica", 11))
        style.configure("TCombobox", font=("Helvetica", 11))

        # --- Form
        form_frame = tk.Frame(root, bg="#f5f8ff")
        form_frame.pack(pady=10)

        labels = [
            ("Spanner / Size", 0, 0), ("Type", 0, 2), ("Length / mm", 0, 4),
            ("Unit", 1, 0), ("Price", 1, 2), ("Grade", 1, 4),
            ("Thread", 2, 0), ("Place", 2, 2)
        ]

        for text, r, c in labels:
            ttk.Label(form_frame, text=text).grid(row=r, column=c, padx=5, pady=5)

        self.spanner_var = tk.StringVar()
        self.type_var = tk.StringVar()
        self.length_var = tk.StringVar()
        self.unit_var = tk.StringVar()
        self.price_var = tk.StringVar()
        self.grade_var = tk.StringVar()
        self.thread_var = tk.StringVar()
        self.place_var = tk.StringVar()

        ttk.Entry(form_frame, textvariable=self.spanner_var, width=20).grid(row=0, column=1)
        self.type_combo = ttk.Combobox(form_frame, textvariable=self.type_var, values=["bolt", "nut", "washer"], state="readonly", width=18)
        self.type_combo.grid(row=0, column=3)
        self.length_entry = ttk.Entry(form_frame, textvariable=self.length_var, width=20)
        self.length_entry.grid(row=0, column=5)
        ttk.Combobox(form_frame, textvariable=self.unit_var, values=["pcs", "kg", "box", "set"], width=18).grid(row=1, column=1)
        ttk.Entry(form_frame, textvariable=self.price_var, width=20).grid(row=1, column=3)
        self.grade_entry = ttk.Entry(form_frame, textvariable=self.grade_var, width=20)
        self.grade_entry.grid(row=1, column=5)
        ttk.Entry(form_frame, textvariable=self.thread_var, width=20).grid(row=2, column=1)
        ttk.Entry(form_frame, textvariable=self.place_var, width=20).grid(row=2, column=3)

        self.type_combo.bind("<<ComboboxSelected>>", self.on_type_change)

        btn_frame = tk.Frame(form_frame, bg="#f5f8ff")
        btn_frame.grid(row=3, column=0, columnspan=6, pady=15)

        ttk.Button(btn_frame, text="Add Product", command=self.add_product).grid(row=0, column=0, padx=10)
        ttk.Button(btn_frame, text="Update", command=self.update_selected).grid(row=0, column=1, padx=10)
        ttk.Button(btn_frame, text="Delete", command=self.delete_selected).grid(row=0, column=2, padx=10)
        ttk.Button(btn_frame, text="Export Bolts", command=lambda: self.export_to_excel("bolt")).grid(row=1, column=0, padx=10, pady=5)
        ttk.Button(btn_frame, text="Export Nuts", command=lambda: self.export_to_excel("nut")).grid(row=1, column=1, padx=10, pady=5)
        ttk.Button(btn_frame, text="Export Washers", command=lambda: self.export_to_excel("washer")).grid(row=1, column=2, padx=10, pady=5)

        # --- Search Box
        search_frame = tk.Frame(root, bg="#f5f8ff")
        search_frame.pack(anchor='e', padx=10)
        ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT, padx=5)
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self.apply_search_filter)
        self.search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=30)
        self.search_entry.pack(side=tk.LEFT)

        # --- Table
        self.tree = ttk.Treeview(root, columns=self.COLUMNS["default"], show="headings")
        self.set_table_columns_for_type("default")
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        self.full_data = []  # Store all rows for search filtering
        self.load_data()
        self.on_type_change()

        # Shortcut Key Bindings
        self.root.bind('<Return>', lambda event: self.add_product())
        self.root.bind('<Delete>', lambda event: self.delete_selected())
        self.root.bind('<Control-u>', lambda event: self.update_selected())

    def set_table_columns_for_type(self, type_):
        cols = self.COLUMNS.get(type_, self.COLUMNS["default"])
        self.tree["columns"] = cols
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=110, anchor=tk.CENTER)

    def on_type_change(self, event=None):
        type_value = self.type_var.get().lower()
        self.set_table_columns_for_type(type_value if type_value else "default")
        if type_value == "bolt":
            self.length_entry.grid()
            self.grade_entry.grid()
        elif type_value == "nut":
            self.length_entry.grid_remove()
            self.grade_entry.grid()
        elif type_value == "washer":
            self.length_entry.grid_remove()
            self.grade_entry.grid_remove()
        else:
            self.length_entry.grid()
            self.grade_entry.grid()
        self.load_data(type_value if type_value else None)

    def load_data(self, type_=None):
        for row in self.tree.get_children():
            self.tree.delete(row)
        products = get_products_by_type(type_) if type_ else get_all_products()
        self.full_data = products
        cols = self.COLUMNS.get(type_, self.COLUMNS["default"])
        for product in products:
            self.tree.insert("", tk.END, values=self.map_product_to_columns(product, cols))

    def map_product_to_columns(self, product, cols):
        id, name, type_, size, unit, price, grade, thread, place = product
        data_map = {
            "ID": id,
            "Spanner / Size": name,
            "Type": type_,
            "Length / mm": size or "",
            "Unit": unit,
            "Price": price,
            "Grade": grade or "",
            "Thread": thread,
            "Place": place or "",
            "Quality": thread  # For washer
        }
        return tuple(data_map.get(col, "") for col in cols)

    def add_product(self):
        try:
            name = self.spanner_var.get()
            type_ = self.type_var.get().lower()
            size = self.length_var.get() if type_ != "nut" and type_ != "washer" else ""
            unit = self.unit_var.get()
            price = float(self.price_var.get()) if self.price_var.get() else 0.0
            grade = self.grade_var.get() if type_ != "washer" else ""
            thread = self.thread_var.get()
            place = self.place_var.get()

            if not all([name, type_, unit, thread, place]):
                raise ValueError("Required fields missing.")

            add_product(name, type_, size, unit, price, grade, thread, place)
            self.clear_form()
            self.load_data(type_)
            self.export_to_excel(type_)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def on_select(self, event):
        selected = self.tree.selection()
        if selected:
            values = self.tree.item(selected[0], "values")
            type_value = self.type_var.get().lower()
            cols = self.COLUMNS.get(type_value, self.COLUMNS["default"])
            field_map = dict(zip(cols, values))
            self.selected_id = field_map.get("ID")
            self.spanner_var.set(field_map.get("Spanner / Size", ""))
            self.type_var.set(field_map.get("Type", ""))
            self.length_var.set(field_map.get("Length / mm", ""))
            self.unit_var.set(field_map.get("Unit", ""))
            self.price_var.set(field_map.get("Price", ""))
            self.grade_var.set(field_map.get("Grade", ""))
            self.thread_var.set(field_map.get("Thread", "") or field_map.get("Quality", ""))
            self.place_var.set(field_map.get("Place", ""))
            self.on_type_change()

    def update_selected(self):
        try:
            if not hasattr(self, 'selected_id'):
                raise ValueError("No product selected.")

            name = self.spanner_var.get()
            type_ = self.type_var.get().lower()
            size = self.length_var.get() if type_ != "nut" and type_ != "washer" else ""
            unit = self.unit_var.get()
            price = float(self.price_var.get()) if self.price_var.get() else 0.0
            grade = self.grade_var.get() if type_ != "washer" else ""
            thread = self.thread_var.get()
            place = self.place_var.get()

            if not all([name, type_, unit, thread, place]):
                raise ValueError("Required fields missing.")

            update_product(self.selected_id, name, type_, size, unit, price, grade, thread, place)
            self.clear_form()
            self.load_data(type_)
            self.export_to_excel(type_)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def delete_selected(self):
        try:
            if not hasattr(self, 'selected_id'):
                raise ValueError("No product selected.")
            type_ = self.type_var.get().lower()
            delete_product(self.selected_id, type_)
            self.clear_form()
            self.load_data(type_)
            self.export_to_excel(type_)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def clear_form(self):
        for var in [self.spanner_var, self.type_var, self.length_var, self.unit_var,
                    self.price_var, self.grade_var, self.thread_var, self.place_var]:
            var.set("")
        if hasattr(self, 'selected_id'):
            del self.selected_id
        self.on_type_change()

    def export_to_excel(self, type_):
        products = get_products_by_type(type_)
        if not products:
            messagebox.showinfo("Export", f"No {type_}s to export.")
            return

        folder = "Hexagon"
        os.makedirs(folder, exist_ok=True)

        file_name = os.path.join(folder, f"{type_}.xlsx")
        headers = list(self.COLUMNS[type_])
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = type_.capitalize() + "s"

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row_num, product in enumerate(products, start=2):
            data = self.map_product_to_columns(product, headers)
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value).alignment = Alignment(horizontal='center')

        try:
            wb.save(file_name)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def apply_search_filter(self, *args):
        keyword = self.search_var.get().lower()
        for row in self.tree.get_children():
            self.tree.delete(row)
        type_value = self.type_var.get().lower()
        cols = self.COLUMNS.get(type_value, self.COLUMNS["default"])
        for product in self.full_data:
            values = self.map_product_to_columns(product, cols)
            if any(keyword in str(val).lower() for val in values):
                self.tree.insert("", tk.END, values=values)


if __name__ == "__main__":
    root = tk.Tk()
    app = ProductApp(root)
    root.mainloop()
